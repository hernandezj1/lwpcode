'''
Convert the separated excel sheets of each indivdual into a formatted JS timeline template so they can de displayed

1. Copy the template row into a renamed file Pedagogue_timeline.xlsx
2. Make the title slide - LWP picture, name of pedagogue and alma mater( column K) if no alma mater check for their first university of record( column F)
3. Go through each row - if row has pub or pres use that as ttiel, populate date, viaf, citation for a new slide, create a row and append into the new
4. if no pub or pres in a row this is a role slide popualte acoordingly, may try statements
5. Create a Contributor slide at the end

'''

'''
Edits made on April May
1. Using the new master dataset which added an extra column topic (LOC) on W where wiki link was previously
Therefore updating all fucntions to reflect this new extraction that must be performed - done
2. Adding titles to the extracted data so they appear on the text itself omn the slide and its not all just together. exactly how Alma mater is already formatted - done
3. Changing all new lines into Breaks - done
3. Need to add new LWP logo

'''

import os
from openpyxl import load_workbook
import openpyxl
from openpyxl import Workbook
import pandas as pd
infolder='individuals'


def template(infile:str,namefile:str):
    outfolder='templates'

    inputworkbook = load_workbook(filename=infile)
    sheet=inputworkbook.active
    titlerow=sheet[1]
    outfilename= namefile[:-5]+'_template.xlsx'
    rename= Workbook()
    outsheet=rename.active
    outsheet.append([cell.value for cell in titlerow])
    rename.save(os.path.join(outfolder,outfilename))


def titleslide(infile:str):

    inputworkbook = load_workbook(filename= (os.path.join('templates',infile)))
    sheet=inputworkbook.active
    name= infile[:-14]
    sheet["J2"] = name
    sheet["P2"] = 'title'
    # Picture media for LWP logo
    sheet["L2"] = 'https://raw.githubusercontent.com/hernandezj1/LWP/main/LWPlogo.png'
    sheet["R2"] = '#3A3A3A' # background to the LWP logo
    inputworkbook.save(os.path.join('templates',infile))


def pubslides(infile:str): 

    # these three lines open the indivudal file
    inputworkbook = load_workbook(filename= (os.path.join('individuals',infile)))
    sheet=inputworkbook.active
    lastrow= sheet.max_row 

    outfile= infile[:-5]+'_template.xlsx'
    outputworkbook = load_workbook(filename= (os.path.join('templates',outfile)))
    outsheet=outputworkbook.active
    


    for i in range(2,lastrow+1): 
        pub= 'M'+str(i)
        value_pub= sheet[pub].value


        if value_pub is not None: 
        # extraction of pub_date- O, pub_type- N , citation- P , & (archive_uri - R or world_cat- Z)
            pub_date= 'O'+str(i)    
            pub_type= 'N'+str(i)  
            citation= 'P'+str(i) 
            archive_uri= 'R' + str(i)    
            world_cat= 'Z' + str(i) 
            viaf ='D'+str(i)
            role= 'C'+str(i)
        
            value_pub_date= sheet[pub_date].value
            value_pub_type= sheet[pub_type].value
            value_citation= sheet[citation].value
            value_archive= sheet[archive_uri].value
            value_world = sheet[world_cat].value
            value_viaf=sheet[viaf].value
            value_role= sheet[role].value

            if value_archive is not None: # decide if archive or world cat
                link=value_archive
            else: 
                link= value_world 
            
            empty_values=['no entry','No entry', 'N/A', None, 'None', ' ', '']

            if value_viaf not in empty_values: 
                viaf_text=value_viaf
            else: 
                viaf_text= 'No record'
            
            text= 'Role: '+str(value_role)+'<br> Publication type: '+ str(value_pub_type)+'<br> Citation: '+ str(value_citation)+ '<br> VIAF entry: '+ str(viaf_text)+'<br> Reference link: '+str(link) # text to fill the slide with the facts and links of the publication
            media='https://raw.githubusercontent.com/hernandezj1/LWP/main/pubmedia.jpg'
            slide_data=[str(value_pub_date),'' ,'' ,'',str(value_pub_date),'','' ,'','',value_pub, text, media, 'phdstudies.com','','','','','#3A3A3A']

            outsheet.append(slide_data)
            for cell in outsheet[outsheet.max_row]: # makes it possible for their to be multiple lines in one cell for the text
                cell.alignment = openpyxl.styles.Alignment(wrapText=True)
            
            # create the list that will be appended to the output worksheet as a new slide
            outputworkbook.save(os.path.join('templates',outfile))
       

def presslides(infile:str): 

    # these three lines open the indivudal file
    inputworkbook = load_workbook(filename= (os.path.join('individuals',infile)))
    sheet=inputworkbook.active
    lastrow= sheet.max_row 

    outfile= infile[:-5]+'_template.xlsx'
    outputworkbook = load_workbook(filename= (os.path.join('templates',outfile)))
    outsheet=outputworkbook.active
    


    for i in range(2,lastrow+1): 
        pres= 'L'+str(i)
        value_pres= sheet[pres].value

        if value_pres is not None: 
        # extraction of pres_date=date range- H, organization- G , citation- P 
            pres_date= 'H'+str(i)     
            citation= 'P'+str(i) 
            org= 'G' + str(i)    


            value_pres_date= sheet[pres_date].value
            value_citation= sheet[citation].value
            value_org= sheet[org].value

            # cleaning up the value of pres date
            value_pres_date=value_pres_date.strip # removes all spaces from the date range

            value_pres_date =str(value_pres_date)
            
            if '-' in value_pres_date:
                start_year = value_pres_date[:4]
                end_year = value_pres_date[-4:]
            else: 
                start_year=value_pres_date
                end_year=value_pres_date

            text= 'Organizations: '+ str(value_org)+'<br> Citation: '+ str(value_citation) # text to fill the slide with the facts and links of the publication
            media='https://raw.githubusercontent.com/hernandezj1/LWP/main/presimage.jpg'
            slide_data=[str(start_year),'' ,'' ,'',str(end_year),'','' ,'','',value_pres, text, media, 'westernventures.com','','','','','#3A3A3A']

            outsheet.append(slide_data)
            for cell in outsheet[outsheet.max_row]: # makes it possible for there to be multiple lines in one cell for the text
                cell.alignment = openpyxl.styles.Alignment(wrapText=True)
            
            # create the list that will be appended to the output worksheet as a new slide
            outputworkbook.save(os.path.join('templates',outfile))


def roleslides(infile:str):

    # these three lines open the indivudal file
    inputworkbook = load_workbook(filename= (os.path.join('individuals',infile)))
    sheet=inputworkbook.active
    lastrow= sheet.max_row 

    outfile= infile[:-5]+'_template.xlsx'
    outputworkbook = load_workbook(filename= (os.path.join('templates',outfile)))
    outsheet=outputworkbook.active

    for i in range(2,lastrow+1): 
        role= 'C'+str(i)
        value_role= sheet[role].value
    
        ignored_roles=['author','co-author','editor','respondent','narrator','narrator-subject', None] # Roles that already have slides 

        if value_role not in ignored_roles: 

            role_date='H'+str(i) 
            univ= 'F' +str(i) 
            orgaf = 'I' +str(i) 
            org = 'G' + str(i)
            political = 'J' +str(i)
            wiki = 'X' +str(i)
            db = 'Y' +str(i)
            worldcat = 'Z' +str(i)
            citation= 'P' + str(i)
            archive= 'R'+ str(i)

            # Selecting which org cell


            date_value= sheet[role_date].value
            univ_value= sheet[univ].value
            orgaf_value= sheet[orgaf].value
            org_value= sheet[org].value
            pol_value= sheet[political].value
            wiki_value= sheet[wiki].value
            db_value= sheet[db].value
            worldcat_value= sheet[worldcat].value
            citation_value= sheet[citation].value
            archive_value= sheet[archive].value

             # Selecting which org cell checks affiliation first if not defaults to the org column
            
            if orgaf_value is None: 
                org_value= org_value
            else: 
                org_value=orgaf_value

            empty_values=['no entry','No entry', 'N/A', None, 'None', ' ', '']
            # Creating the full citation
            fullcitation=''
            if citation_value not in empty_values: 
                fullcitation += 'Citation:'+ str(citation_value) + '<br>'
            if archive_value not in empty_values: 
                fullcitation += 'Archive link:'+ str(archive_value)+ '<br>'

            # deciding between the links


            if worldcat_value not in empty_values: 
                link= worldcat_value
            
            elif wiki_value not in empty_values: 
                link =wiki_value

            else: 
                link= db_value

            # Creating the text that will be added to the slide, check if org and pol are None but still add them if they are present
            text = ''
            if univ_value is not None:
                text += 'Institution: ' + str(univ_value) + '<br>'
            if org_value is not None:
                text += 'Organization(s): ' + str(org_value) + '<br>'
            if pol_value is not None:
                text += 'Political affiliation(s): '+ str(pol_value) + '<br>'
            if link not in empty_values: 
                text += 'Reference link: '+ str(link) + '<br>'
            text += str(fullcitation) # adding  archive URI and citation text 

            if date_value is not None: # if there is an actual date
                
                # cleaning up the value of date
                date_value=str(date_value)
                date_value=date_value.strip() # removes all spaces from the date range

                if '-' in date_value:
                    start_year = date_value[:4]
                    end_year = date_value[-4:]
                else: 
                    start_year=date_value
                    end_year=date_value

                media='https://raw.githubusercontent.com/hernandezj1/LWP/main/roleimage.jpg'
                slide_data=[str(start_year),'' ,'' ,'',str(end_year),'','' ,'','',value_role, text, media, 'irisreading.com','','','','','#3A3A3A']

                outsheet.append(slide_data)
                for cell in outsheet[outsheet.max_row]: # makes it possible for their to be multiple lines in one cell for the text
                    cell.alignment = openpyxl.styles.Alignment(wrapText=True)
            
                # create the list that will be appended to the output worksheet as a new slide
                outputworkbook.save(os.path.join('templates',outfile))               
            
            else: # here we will append the data to text of the title slide
                complete_text= value_role+ ' ' + text # Adds the role name because the previous text accounts for the title in slide
                title_text_value= outsheet['K2'].value        #load existing text
                if title_text_value is not None:
                    newtext = title_text_value+'<br>'+complete_text
                else: 
                    newtext = complete_text
                outsheet['K2']= newtext
                outputworkbook.save(os.path.join('templates',outfile))   
        
def almamater(infile:str): 
    
    # these three lines open the indivudal file
    inputworkbook = load_workbook(filename= (os.path.join('individuals',infile)))
    sheet=inputworkbook.active
    lastrow= sheet.max_row 

    outfile= infile[:-5]+'_template.xlsx'
    outputworkbook = load_workbook(filename= (os.path.join('templates',outfile)))
    outsheet=outputworkbook.active
    almamater_value= None
    # Goes thorugh the alma mater column until it find the first entry, all entries are identical
    for i in range(2,lastrow+1): 
        am= 'K'+str(i)
        value_am= sheet[am].value
        if value_am is not None:
            almamater_value=value_am
            break
        
    if almamater_value is not None: 

        almamater_value= almamater_value.replace(';',' <br> ')
        almamatertext= 'Alma Mater: <br>'+ almamater_value
        title_text_value= outsheet['K2'].value        #load existing text
        if title_text_value is not None:
            newtext = almamatertext+ '<br> '+ title_text_value
        else: 
            newtext = almamatertext
        outsheet['K2']= newtext
        outputworkbook.save(os.path.join('templates',outfile))      

def reorder(infile:str): 
    data=pd.read_excel(os.path.join('templates',infile))
    data = data.sort_values(by='Year', na_position='first')
    data.to_excel(os.path.join('templates',infile), index = False)


    




for file in os.listdir(infolder): # adds templates header

    template('template.xlsx',file)

for file in os.listdir('templates'): # adds a title slide
    titleslide(file)

for file in os.listdir(infolder): # add all publication slides 
    pubslides(file)

for file in os.listdir(infolder): # add all presentation slides 
    presslides(file)

for file in os.listdir(infolder): # add all presentation slides 
    roleslides(file)

for file in os.listdir(infolder): # adds Alma mater to the title slide
    almamater(file)

for file in os.listdir('templates'): # Reorders by beginning year
    reorder(file)