"""
First step of the LWP deployment process
Dividing the master dataset into individual sheets each with a single name

For Clarissa "Clara" Higgins rememebr that the "" where eliminated for processing need to be added individually
the blank space was elimiated as well as the name that Tarez had as a M?
"""
import os
from openpyxl import load_workbook
from openpyxl import Workbook
inputworkbook = load_workbook(filename="Master Datasets.xlsx")

Mastersheet=inputworkbook.active


for i in range(2,498): 
    currentindex='B'+str(i)
    currentname= Mastersheet[currentindex].value
    filename= currentname +'.xlsx'
    Fieldsrow=Mastersheet[1]
    row=Mastersheet[i]


    if os.path.exists(filename) != True: # if file does not exist create the file and add the new row
        output= Workbook()
        sheet= output.active
        sheet.append([cell.value for cell in Fieldsrow])
        sheet.append([cell.value for cell in row])
        output.save(filename=filename)
    else:                               # if file does exist append the row to the existing file
        outputworkbook = load_workbook(filename=filename)
        sheet= outputworkbook.active
        sheet.append([cell.value for cell in row])
        outputworkbook.save(filename=filename)
